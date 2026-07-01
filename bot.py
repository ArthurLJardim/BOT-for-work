from __future__ import annotations

import argparse
import calendar
import csv
import hmac
import imaplib
import io
import json
import logging
import os
import re
import hashlib
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import shutil
import smtplib
import ssl
import time
import unicodedata
import urllib.parse
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from email import policy
from email.message import EmailMessage, Message
from email.parser import BytesParser
from email.utils import parseaddr
from functools import lru_cache
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = SCRIPT_DIR / "config.json"
BANCOS_PATH = SCRIPT_DIR / "bancos.json"
EMAIL_CONFIG_PATH = SCRIPT_DIR / "email_config.json"
DEFAULT_PASTA_BASE = Path(r"F:\TRABALHO\SIMPLES NACIONAL 2026")
DEFAULT_PLANILHA = Path(r"F:\TRABALHO\BOT\TESTE  SIMPLES  NACIONAL.xlsx")
DEFAULT_ANO = 2026
DEFAULT_WHATSAPP_INBOX = SCRIPT_DIR / "entrada_whatsapp"
DEFAULT_ASSISTANT_CHARGES = SCRIPT_DIR / "relatorios" / "cobrancas_assistente.csv"
DEFAULT_APPROVAL_CONTACT = "BOT"
DEFAULT_WEBHOOK_TOKEN = "esperanto_bot_2026"
DEFAULT_WEBHOOK_LOG = SCRIPT_DIR / "relatorios" / "whatsapp_webhook_eventos.jsonl"
ATTACHMENT_EXTENSIONS = {".pdf", ".ofx", ".xls", ".xlsx", ".csv", ".png", ".jpg", ".jpeg", ".zip"}

logging.getLogger("pypdf").setLevel(logging.ERROR)
logging.getLogger("PyPDF2").setLevel(logging.ERROR)

HEADER_ALIASES = {
    "codigo": {"codigo", "cod", "codigo empresa"},
    "empresa": {"empresa", "nome", "cliente", "razao social"},
    "faltando": {"faltando", "faltantes", "pendente", "pendencias"},
}

STATEMENT_EXTENSIONS = {".pdf", ".ofx"}
STATEMENT_KEYWORDS = {
    "extrato",
    "pdf",
    "ofx",
    "janeiro",
    "fevereiro",
    "marco",
    "abril",
    "maio",
    "junho",
    "julho",
    "agosto",
    "setembro",
    "outubro",
    "novembro",
    "dezembro",
}
MONTH_NAME_HINTS = {
    "janeiro",
    "jan",
    "fevereiro",
    "fev",
    "marco",
    "mar o",
    "mar oo",
    "abril",
    "abr",
    "maio",
    "mai",
    "junho",
    "jun",
    "julho",
    "jul",
    "agosto",
    "ago",
    "setembro",
    "set",
    "outubro",
    "out",
    "novembro",
    "nov",
    "dezembro",
    "dez",
}
DEFAULT_BANK_ALIASES = {
    "brb": "BRB",
    "banco de brasilia": "BRB",
    "bb": "BB",
    "banco do brasil": "BB",
    "itau": "ITAU",
    "itaú": "ITAU",
    "nubank": "NUBANK",
    "nu": "NUBANK",
    "sicoob": "SICOOB",
    "sicredi": "SICREDI",
    "bradesco": "BRADESCO",
    "santander": "SANTANDER",
    "caixa economica": "CAIXA",
    "caixa economica federal": "CAIXA",
    "cef": "CAIXA",
    "inter": "INTER",
}
BANK_ID_ALIASES = {
    "001": "BB",
    "1": "BB",
    "070": "BRB",
    "70": "BRB",
    "237": "BRADESCO",
    "756": "SICOOB",
    "748": "SICREDI",
    "104": "CAIXA",
    "033": "SANTANDER",
    "33": "SANTANDER",
    "341": "ITAU",
    "260": "NUBANK",
    "336": "C6",
    "197": "STONE",
}


@dataclass(frozen=True)
class Config:
    pasta_base: Path
    planilha: Path
    ano: int


@dataclass(frozen=True)
class EmailConfig:
    imap_host: str
    imap_port: int
    smtp_host: str
    smtp_port: int
    usuario: str
    senha: str
    pasta_entrada: str
    marcar_como_lido: bool


def load_config(path: Path = CONFIG_PATH) -> Config:
    if not path.exists():
        return Config(DEFAULT_PASTA_BASE, DEFAULT_PLANILHA, DEFAULT_ANO)

    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)

    return Config(
        pasta_base=Path(data.get("pasta_base", DEFAULT_PASTA_BASE)),
        planilha=Path(data.get("planilha", DEFAULT_PLANILHA)),
        ano=int(data.get("ano", DEFAULT_ANO)),
    )


def load_email_config(path: Path = EMAIL_CONFIG_PATH) -> EmailConfig:
    if not path.exists():
        raise FileNotFoundError(
            f"Configuracao de e-mail nao encontrada: {path}. "
            "Crie esse arquivo a partir do email_config.exemplo.json."
        )

    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)

    password = data.get("senha") or os.environ.get(data.get("senha_env", "BOT_EMAIL_PASSWORD"), "")
    if not password:
        raise ValueError("Senha do e-mail nao configurada. Use 'senha' no email_config.json ou a variavel de ambiente.")

    return EmailConfig(
        imap_host=str(data["imap_host"]),
        imap_port=int(data.get("imap_port", 993)),
        smtp_host=str(data["smtp_host"]),
        smtp_port=int(data.get("smtp_port", 587)),
        usuario=str(data["usuario"]),
        senha=str(password),
        pasta_entrada=str(data.get("pasta_entrada", "INBOX")),
        marcar_como_lido=bool(data.get("marcar_como_lido", False)),
    )


def load_bank_aliases(path: Path = BANCOS_PATH) -> dict[str, str]:
    if not path.exists():
        return DEFAULT_BANK_ALIASES

    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)

    aliases: dict[str, str] = {}
    for bank, names in data.items():
        if bank not in {"CAIXA", "C6"}:
            aliases[bank] = bank
            aliases[bank.casefold()] = bank
        for name in names:
            alias_key = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode("ascii").casefold()
            alias_key = re.sub(r"[^a-z0-9]+", " ", alias_key).strip()
            if bank == "CAIXA" and alias_key == "cef":
                continue
            aliases[str(name)] = bank
            aliases[alias_key] = bank
    return aliases


CONFIG = load_config()
BANK_ALIASES = load_bank_aliases()
PASTA_BASE = CONFIG.pasta_base
PLANILHA = CONFIG.planilha
ANO = CONFIG.ano


@dataclass(frozen=True)
class Columns:
    header_row: int
    codigo: int
    empresa: int
    faltando: int


@dataclass(frozen=True)
class CompanyRecord:
    codigo: object
    empresa: object
    pasta: Path | None
    codigo_text: str
    empresa_text: str
    empresa_tokens: tuple[str, ...]


def normalize(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"\s+", " ", text).strip().casefold()
    return text


def normalize_for_match(value: object) -> str:
    text = normalize(value)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def current_default_end_month(year: int, include_current: bool) -> int:
    today = date.today()
    if year < today.year:
        return 12
    if year > today.year:
        raise ValueError("Informe --meses ao conferir um ano futuro.")
    return today.month if include_current else today.month - 1


def parse_months(value: str | None, year: int, include_current: bool) -> list[int]:
    if value:
        months: set[int] = set()
        for part in re.split(r"[,\s;]+", value.strip()):
            if not part:
                continue
            if "-" in part:
                start_text, end_text = part.split("-", 1)
                start, end = int(start_text), int(end_text)
                months.update(range(start, end + 1))
            else:
                months.add(int(part))
        ordered = sorted(months)
    else:
        end_month = current_default_end_month(year, include_current)
        if end_month < 1:
            raise ValueError(
                "Ainda nao ha mes fechado para conferir. Informe --meses ou use --incluir-mes-atual."
            )
        ordered = list(range(1, end_month + 1))

    invalid = [month for month in ordered if month < 1 or month > 12]
    if invalid:
        raise ValueError(f"Meses invalidos: {invalid}. Use valores de 1 a 12.")

    if not ordered:
        raise ValueError(f"Nenhum mes para conferir no ano {year}.")
    return ordered


def month_folder_names(months: Iterable[int], year: int) -> dict[int, str]:
    return {month: f"{month:02d} - {year}" for month in months}


def month_label(month: int) -> str:
    return f"{month:02d}"


def find_columns(sheet: Worksheet, header_row: int | None) -> Columns:
    rows_to_scan = [header_row] if header_row else range(1, min(sheet.max_row, 20) + 1)

    for row in rows_to_scan:
        found: dict[str, int] = {}
        for cell in sheet[row]:
            header = normalize(cell.value)
            for key, aliases in HEADER_ALIASES.items():
                if header in aliases:
                    found[key] = cell.column

        if {"codigo", "empresa", "faltando"}.issubset(found):
            return Columns(
                header_row=row,
                codigo=found["codigo"],
                empresa=found["empresa"],
                faltando=found["faltando"],
            )

    raise ValueError(
        "Nao encontrei as colunas CODIGO, EMPRESA e FALTANDO. "
        "Confira se os cabecalhos existem na planilha."
    )


def find_optional_column(sheet: Worksheet, header_row: int, aliases: set[str]) -> int | None:
    for cell in sheet[header_row]:
        if normalize(cell.value) in aliases:
            return cell.column
    return None


def list_company_folders(base_folder: Path) -> list[Path]:
    if not base_folder.exists():
        raise FileNotFoundError(f"Pasta base nao encontrada: {base_folder}")
    return [path for path in base_folder.iterdir() if path.is_dir()]


def find_company_folder(company_folders: list[Path], codigo: object, empresa: object) -> Path | None:
    code_text = str(codigo).strip() if codigo is not None else ""
    code_text = re.sub(r"\.0$", "", code_text)
    company_text = normalize_for_match(empresa)

    if code_text:
        code_pattern = re.compile(rf"^\s*{re.escape(code_text)}(?:\D|$)")
        for folder in company_folders:
            if code_pattern.search(folder.name):
                return folder

    if company_text:
        for folder in company_folders:
            if company_text in normalize_for_match(folder.name):
                return folder

    return None


def month_folder(company_folder: Path, month: int, year: int) -> Path | None:
    expected_name = normalize(f"{month:02d} - {year}")
    flexible_pattern = re.compile(rf"^0?{month}\s*-\s*{year}$")
    month_only_pattern = re.compile(rf"^0?{month}(?:\D|$)")
    fallback = None

    for child in company_folder.iterdir():
        if not child.is_dir():
            continue
        child_name = normalize(child.name)
        if child_name == expected_name or flexible_pattern.match(child_name):
            return child
        if fallback is None and month_only_pattern.match(child_name):
            fallback = child

    return fallback


def months_from_file_name(file_name: str, year: int, expected_months: set[int]) -> set[int]:
    text = normalize(file_name)
    found: set[int] = set()

    range_pattern = re.compile(
        rf"\b(0?[1-9]|1[0-2])\s*(?:a|ate|ao|e)\s*(0?[1-9]|1[0-2])\s*[-_/ ]+\s*{year}\b"
    )
    for match in range_pattern.finditer(text):
        start, end = int(match.group(1)), int(match.group(2))
        if start <= end:
            found.update(range(start, end + 1))

    range_without_year_pattern = re.compile(r"\b(0?[1-9]|1[0-2])\s*(?:-|a|ate|ao|e)\s*(0?[1-9]|1[0-2])\b")
    for match in range_without_year_pattern.finditer(text):
        start, end = int(match.group(1)), int(match.group(2))
        if start <= end:
            found.update(range(start, end + 1))

    single_pattern = re.compile(rf"\b(0?[1-9]|1[0-2])\s*[-_/ ]+\s*{year}\b")
    for match in single_pattern.finditer(text):
        found.add(int(match.group(1)))

    return found & expected_months


def known_bank_in_text(text: str) -> str | None:
    normalized_text = normalize_for_match(text)
    compact_text = normalized_text.replace(" ", "")
    words = set(normalized_text.split())

    if "extratobb" in compact_text or "bancodobrasil" in compact_text:
        return "BB"

    for alias, bank in sorted(BANK_ALIASES.items(), key=lambda item: len(item[0]), reverse=True):
        normalized_alias = normalize_for_match(alias)
        if " " in normalized_alias:
            if normalized_alias in normalized_text:
                return bank
            continue
        if normalized_alias in words:
            return bank

    return None


def known_bank_in_file_name(text: str) -> str | None:
    normalized_text = normalize_for_match(text)
    if re.search(r"\bcef\b", normalized_text):
        return "CAIXA"
    return known_bank_in_text(normalized_text)


def has_month_hint(text: str) -> bool:
    padded = f" {text} "
    if re.search(r"\b(0?[1-9]|1[0-2])\b", text):
        return True
    return any(f" {month_name} " in padded for month_name in MONTH_NAME_HINTS)


def file_text_for_bank_detection(file: Path) -> str:
    suffix = file.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf_text(file, max_pages=2)
    if suffix == ".ofx":
        return read_text_file(file, max_chars=12000)
    return ""


def read_text_file(file: Path, max_chars: int = 12000) -> str:
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return file.read_text(encoding=encoding, errors="ignore")[:max_chars]
        except Exception:
            continue
    return ""


def bank_from_ofx(file: Path) -> str | None:
    text = read_text_file(file, max_chars=6000)
    normalized_text = normalize_for_match(text)

    bank_id_match = re.search(r"<\s*bankid\s*>\s*0*(\d+)", text, flags=re.IGNORECASE)
    if bank_id_match:
        bank = BANK_ID_ALIASES.get(bank_id_match.group(1))
        if bank:
            return bank

    fid_match = re.search(r"<\s*fid\s*>\s*0*(\d+)", text, flags=re.IGNORECASE)
    if fid_match:
        bank = BANK_ID_ALIASES.get(fid_match.group(1))
        if bank:
            return bank

    org_match = re.search(r"<\s*org\s*>\s*([^<\r\n]+)", text, flags=re.IGNORECASE)
    if org_match:
        org_text = normalize_for_match(org_match.group(1))
        if "banco cooperativo do brasil" in org_text:
            return "SICOOB"
        bank = known_bank_in_text(org_text)
        if bank:
            return bank

    return known_bank_in_text(normalized_text[:500])


def account_ids_from_text(text: str) -> set[str]:
    normalized = normalize(text)
    accounts = set()

    for match in re.finditer(r"<\s*acctid\s*>\s*([0-9][0-9\.\- ]{3,})", normalized):
        digits = re.sub(r"\D", "", match.group(1))
        if len(digits) >= 4:
            accounts.add(digits)

    for match in re.finditer(r"(?:conta corrente|acctid|conta)\s*[:\-]?\s*([0-9][0-9\.\- ]{3,})", normalized):
        digits = re.sub(r"\D", "", match.group(1))
        if len(digits) >= 4:
            accounts.add(digits)

    return accounts


def account_ids_from_file(file: Path) -> set[str]:
    text = file_text_for_bank_detection(file)
    return account_ids_from_text(text)


def cents_from_number_text(value: str) -> int | None:
    number = parse_brazilian_number(value)
    if number is None:
        return None
    return int(round(number * 100))


@lru_cache(maxsize=512)
def ofx_transaction_signatures(file: Path) -> frozenset[tuple[str, int]]:
    text = read_text_file(file, max_chars=400000)
    signatures: set[tuple[str, int]] = set()

    for block in re.findall(r"<\s*stmttrn\s*>(.*?)(?=<\s*stmttrn\s*>|</\s*banktranlist\s*>|$)", text, flags=re.IGNORECASE | re.DOTALL):
        date_match = re.search(r"<\s*dtposted\s*>\s*(\d{8})", block, flags=re.IGNORECASE)
        amount_match = re.search(r"<\s*trnamt\s*>\s*(-?\d+(?:[.,]\d+)?)", block, flags=re.IGNORECASE)
        if not date_match or not amount_match:
            continue
        amount = amount_match.group(1).replace(",", ".")
        try:
            cents = int(round(float(amount) * 100))
        except ValueError:
            continue
        signatures.add((date_match.group(1)[:8], abs(cents)))

    return frozenset(signatures)


@lru_cache(maxsize=512)
def pdf_transaction_signatures(file: Path) -> frozenset[tuple[str, int]]:
    text = extract_pdf_text(file, max_pages=8)
    signatures: set[tuple[str, int]] = set()
    if not text:
        return frozenset()

    for raw_line in text.splitlines():
        line = normalize(raw_line)
        dates = list(re.finditer(r"\b(\d{1,2})/(\d{1,2})(?:/(\d{4}))?\b", line))
        if not dates:
            continue
        amounts = re.findall(r"-?\d[\d\.\s]*,\d{2}", line)
        if not amounts:
            continue

        day, month, year = dates[0].group(1), dates[0].group(2), dates[0].group(3)
        if year:
            date_key = f"{int(year):04d}{int(month):02d}{int(day):02d}"
        else:
            date_key = f"{int(month):02d}{int(day):02d}"

        for amount_text in amounts:
            cents = cents_from_number_text(amount_text)
            if cents is not None and cents != 0:
                signatures.add((date_key, abs(cents)))

    return frozenset(signatures)


def matching_transaction_count(pdf_signatures: frozenset[tuple[str, int]], ofx_signatures: frozenset[tuple[str, int]]) -> int:
    if not pdf_signatures or not ofx_signatures:
        return 0

    full_matches = pdf_signatures & ofx_signatures
    if full_matches:
        return len(full_matches)

    pdf_short = {(date_key[-4:], cents) for date_key, cents in pdf_signatures}
    ofx_short = {(date_key[-4:], cents) for date_key, cents in ofx_signatures}
    return len(pdf_short & ofx_short)


def sibling_bank_by_account(file: Path) -> str | None:
    current_accounts = account_ids_from_file(file)
    if not current_accounts:
        return None

    for sibling in file.parent.iterdir():
        if sibling == file or not sibling.is_file() or sibling.suffix.lower() not in STATEMENT_EXTENSIONS:
            continue
        sibling_bank = bank_from_statement_file_by_name_or_content(sibling)
        if not sibling_bank:
            continue
        sibling_accounts = account_ids_from_file(sibling)
        if current_accounts & sibling_accounts:
            return sibling_bank

    return None


def sibling_bank_by_unique_ofx(file: Path) -> str | None:
    banks = set()
    for sibling in file.parent.iterdir():
        if sibling == file or not sibling.is_file() or sibling.suffix.lower() != ".ofx":
            continue
        sibling_bank = bank_from_statement_file_by_name_or_content(sibling)
        if sibling_bank:
            banks.add(sibling_bank)
    if len(banks) == 1:
        return next(iter(banks))
    return None


def sibling_bank_by_transactions(file: Path) -> str | None:
    text = normalize_for_match(file.stem)
    compact_text = text.replace(" ", "")
    if "contasapagar" in compact_text or "contaspagas" in compact_text or "relatorio" in text or "balancete" in text:
        return None

    pdf_signatures = pdf_transaction_signatures(file)
    if not pdf_signatures:
        return None

    best_bank = None
    best_count = 0
    tied = False
    for sibling in file.parent.iterdir():
        if sibling == file or not sibling.is_file() or sibling.suffix.lower() != ".ofx":
            continue
        bank = bank_from_statement_file_by_name_or_content(sibling)
        if not bank:
            continue
        count = matching_transaction_count(pdf_signatures, ofx_transaction_signatures(sibling))
        if count > best_count:
            best_bank = bank
            best_count = count
            tied = False
        elif count == best_count and count > 0:
            tied = True

    if best_bank and best_count >= 3 and not tied:
        return best_bank
    return None


def sibling_ofx_files_for_bank(file: Path, bank: str) -> list[Path]:
    files = []
    for sibling in file.parent.iterdir():
        if sibling == file or not sibling.is_file() or sibling.suffix.lower() != ".ofx":
            continue
        if bank_from_statement_file_by_name_or_content(sibling) == bank:
            files.append(sibling)
    return files


def pdf_matches_bank_ofx(file: Path, bank: str) -> bool | None:
    pdf_signatures = pdf_transaction_signatures(file)
    if not pdf_signatures:
        return None

    ofx_files = sibling_ofx_files_for_bank(file, bank)
    if not ofx_files:
        return None

    best_count = max(
        matching_transaction_count(pdf_signatures, ofx_transaction_signatures(ofx_file))
        for ofx_file in ofx_files
    )
    return best_count >= 3


def bank_from_statement_file_by_name_or_content(file: Path) -> str | None:
    text = normalize_for_match(file.stem)
    compact_text = text.replace(" ", "")
    if "contaspagas" in compact_text or "contasrecebidas" in compact_text:
        return None

    bank = known_bank_in_file_name(text)
    if bank:
        return bank

    if file.suffix.lower() == ".ofx":
        return bank_from_ofx(file)

    content_text = normalize_for_match(file_text_for_bank_detection(file))
    return known_bank_in_text(content_text)


def bank_from_statement_file(file: Path) -> str | None:
    if file.suffix.lower() not in STATEMENT_EXTENSIONS:
        return None
    if not is_bank_statement_file(file):
        if file.suffix.lower() == ".pdf":
            return sibling_bank_by_transactions(file)
        return None

    if file.suffix.lower() == ".pdf":
        bank = known_bank_in_file_name(file.stem)
        if bank == "CAIXA":
            matches_caixa_ofx = pdf_matches_bank_ofx(file, "CAIXA")
            if matches_caixa_ofx is False:
                bank = None
        if not bank:
            bank = sibling_bank_by_account(file)
        if not bank:
            bank = sibling_bank_by_transactions(file)
        if not bank:
            bank = sibling_bank_by_unique_ofx(file)
        if not bank:
            bank = bank_from_statement_file_by_name_or_content(file)
    else:
        bank = bank_from_statement_file_by_name_or_content(file)
    if not bank:
        return None

    return bank


def is_bank_statement_file(file: Path) -> bool:
    text = normalize_for_match(file.stem)
    compact_text = text.replace(" ", "")
    if any(word in text for word in ("contas a pagar", "contas pagas", "juros pagos")):
        return False
    if any(word in text for word in ("invest", "invt", "investimento")):
        return False
    if "contasapagar" in compact_text or "contaspagas" in compact_text:
        return False
    if file.suffix.lower() == ".ofx":
        return True
    if any(word in text for word in ("extrato", " cc ", " conta corrente")) or text.endswith(" cc"):
        return True
    if re.search(r"\bextr\b", text):
        return True
    return known_bank_in_file_name(text) is not None and has_month_hint(text)


def has_statement_file(month_path: Path, bank: str, extension: str) -> bool:
    return bool(statement_files(month_path, bank, extension))


def statement_files(month_path: Path, bank: str, extension: str | None = None) -> list[Path]:
    files = []
    for file in month_path.iterdir():
        if not file.is_file():
            continue
        if extension and file.suffix.lower() != extension:
            continue
        if bank_from_statement_file(file) == bank:
            files.append(file)
    return files


def has_paid_accounts_report(month_path: Path) -> bool:
    for file in month_path.iterdir():
        if not file.is_file():
            continue

        text = normalize_for_match(file.stem)
        compact_text = text.replace(" ", "")
        if (
            "contas pagas" in text
            or "conta paga" in text
            or "contas pg" in text
            or "conta pg" in text
            or "contaspagas" in compact_text
            or "contapaga" in compact_text
            or "contaspg" in compact_text
            or "contapg" in compact_text
        ):
            return True

        if file.suffix.lower() == ".pdf" and is_paid_accounts_pdf(file):
            return True

    return False


def is_paid_accounts_pdf(file: Path) -> bool:
    text = extract_pdf_text(file, max_pages=2)
    if not text:
        return False

    normalized_text = normalize_for_match(text)
    compact_text = normalized_text.replace(" ", "")

    has_expense_report = (
        "relacao de despesas" in normalized_text
        and "data lancamento" in normalized_text
        and (
            "descricao especifica" in normalized_text
            or ("descricao" in normalized_text and "valor" in normalized_text)
        )
    )
    has_payment_details = (
        ("pagamento" in normalized_text or "despesas" in normalized_text)
        and "valor" in normalized_text
        and ("descricao" in normalized_text or "especifica" in normalized_text)
    )

    return has_expense_report or has_payment_details or "relacaodedespesas" in compact_text


@lru_cache(maxsize=512)
def extract_pdf_text(file: Path, max_pages: int = 2) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            return ""

    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            reader = PdfReader(file)
            pages = reader.pages[:max_pages]
            return "\n".join(page.extract_text() or "" for page in pages)
    except Exception as error:
        print(f"Nao consegui ler o PDF {file}: {error}")
        return ""


def statement_banks_in_month(month_path: Path) -> set[str]:
    banks = set()
    for file in month_path.iterdir():
        if not file.is_file():
            continue
        bank = bank_from_statement_file(file)
        if bank:
            banks.add(bank)
    return banks


def expected_statement_banks(company_folder: Path, month: int, year: int) -> set[str]:
    banks = set()
    for previous_month in range(1, month):
        previous_folder = month_folder(company_folder, previous_month, year)
        if previous_folder:
            banks.update(statement_banks_in_month(previous_folder))
    if not banks:
        banks.update(all_statement_banks(company_folder, year))
    return banks


def all_statement_banks(company_folder: Path, year: int) -> set[str]:
    banks = set()
    for child in company_folder.iterdir():
        if not child.is_dir():
            continue
        child_name = normalize(child.name)
        is_month_folder = re.match(r"^(0?[1-9]|1[0-2])(?:\D|$)", child_name)
        if not is_month_folder and not re.search(rf"\b{year}\b", child_name):
            continue
        banks.update(statement_banks_in_month(child))
    return banks


@lru_cache(maxsize=512)
def first_statement_months_by_bank(company_folder: Path, year: int) -> dict[str, int]:
    first_months: dict[str, int] = {}
    for month in range(1, 13):
        folder = month_folder(company_folder, month, year)
        if not folder:
            continue
        for bank in statement_banks_in_month(folder):
            first_months[bank] = min(month, first_months.get(bank, month))
    return first_months


def should_charge_bank_for_month(company_folder: Path, bank: str, month: int, year: int) -> bool:
    first_months = first_statement_months_by_bank(company_folder, year)
    first_month = first_months.get(bank)
    if first_month is None or month >= first_month:
        return True

    first_folder = month_folder(company_folder, first_month, year)
    if not first_folder:
        return True

    return first_statement_has_previous_balance(first_folder, bank)


def first_statement_has_previous_balance(month_path: Path, bank: str) -> bool:
    for file in statement_files(month_path, bank, ".pdf"):
        balance = previous_balance_from_pdf(file)
        if balance is not None:
            return balance != 0
    return False


@lru_cache(maxsize=512)
def previous_balance_from_pdf(file: Path) -> float | None:
    text = extract_pdf_text(file, max_pages=2)
    if not text:
        return None

    normalized_text = normalize(text)
    match = re.search(r"saldo anterior\s+(-?\d[\d\.\s]*,\d{2}|-?\d+(?:\.\d{2})?)", normalized_text)
    if not match:
        return None

    return parse_brazilian_number(match.group(1))


def parse_brazilian_number(value: str) -> float | None:
    text = value.strip().replace(" ", "")
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def missing_items(company_folder: Path | None, expected: dict[int, str]) -> list[str]:
    if not company_folder or not company_folder.exists():
        return [month_label(month) for month in expected]

    covered_by_file = set()
    expected_months = set(expected)
    year = int(next(iter(expected.values()))[-4:])
    for file in company_folder.iterdir():
        if file.is_file():
            covered_by_file.update(months_from_file_name(file.name, year, expected_months))

    missing = []
    reference_banks = all_statement_banks(company_folder, year)
    for month in expected:
        current_folder = month_folder(company_folder, month, year)
        if not current_folder:
            if reference_banks:
                for bank in sorted(reference_banks):
                    if month not in covered_by_file and should_charge_bank_for_month(company_folder, bank, month, year):
                        missing.append(f"{bank} PDF - {month_label(month)}")
                        missing.append(f"{bank} OFX - {month_label(month)}")
                if month not in covered_by_file:
                    missing.append(f"CP - {month_label(month)}")
            elif month not in covered_by_file:
                missing.append(month_label(month))
            continue

        banks = expected_statement_banks(company_folder, month, year)
        banks.update(
            bank
            for bank in reference_banks
            if should_charge_bank_for_month(company_folder, bank, month, year)
        )
        for bank in sorted(banks):
            if not should_charge_bank_for_month(company_folder, bank, month, year):
                continue
            if not has_statement_file(current_folder, bank, ".pdf"):
                missing.append(f"{bank} PDF - {month_label(month)}")
            if not has_statement_file(current_folder, bank, ".ofx"):
                missing.append(f"{bank} OFX - {month_label(month)}")

        if not has_paid_accounts_report(current_folder):
            missing.append(f"CP - {month_label(month)}")

    return missing


def analyzed_files(company_folder: Path | None, expected: dict[int, str]) -> list[str]:
    if not company_folder or not company_folder.exists():
        return []

    year = int(next(iter(expected.values()))[-4:])
    files = []
    for file in company_folder.iterdir():
        if file.is_file():
            files.append(file.name)

    for month in expected:
        folder = month_folder(company_folder, month, year)
        if not folder:
            continue
        for file in folder.iterdir():
            if file.is_file():
                files.append(f"{folder.name}\\{file.name}")

    return sorted(set(files))


def write_report(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["codigo", "empresa", "canal", "telefone", "email", "pasta", "arquivos_analisados", "faltando"]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def write_charges(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["codigo", "empresa", "canal", "destino", "faltando", "mensagem"]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def row_values(sheet: Worksheet, row: int) -> list[object]:
    return [sheet.cell(row=row, column=column).value for column in range(1, sheet.max_column + 1)]


def extract_emails(values: Iterable[object]) -> list[str]:
    emails = []
    for value in values:
        if value is None:
            continue
        emails.extend(re.findall(r"[\w\.-]+@[\w\.-]+\.\w+", str(value)))
    return sorted(set(emails))


def extract_phones(values: Iterable[object]) -> list[str]:
    phones = []
    for value in values:
        if value is None:
            continue
        text = str(value)
        if "@" in text:
            continue
        for match in re.findall(r"\+?\d[\d\s\-\(\)]{7,}\d", text):
            digits = re.sub(r"\D", "", match)
            if len(digits) >= 10:
                phones.append(digits)
    return sorted(set(phones))


def choose_charge_channel(via: object, phones: list[str], emails: list[str]) -> tuple[str, str]:
    via_text = normalize_for_match(via)
    has_whats = "whats" in via_text
    has_email = "email" in via_text

    if has_email:
        return ("email", emails[0] if emails else "")
    if has_whats:
        return ("whatsapp", phones[0] if phones else "")
    if emails:
        return ("email", emails[0])
    if phones:
        return ("whatsapp", phones[0])
    return ("sem contato", "")


def charge_message(empresa: object, faltando: str, channel: str) -> str:
    faltando_cliente = charge_missing_text(faltando)
    if channel == "whatsapp":
        greeting = "Bom dia" if datetime.now().hour < 12 else "Boa tarde"
        return (
            f"{greeting}, tudo bom? Notamos que alguns documentos ainda não foram enviados, "
            f"sendo eles: {faltando_cliente}"
        )

    return f"Boa tarde\n\nNotamos que alguns documentos ainda não foram enviados, sendo eles: {faltando_cliente}\n\ngentilmente,\nArthur Lopes."


def charge_missing_text(faltando: str) -> str:
    parts = [part.strip() for part in faltando.split(",") if part.strip()]
    formatted = []

    for part in parts:
        cp_match = re.fullmatch(r"CP (.+)", part)
        if cp_match:
            formatted.append(f"relatório de contas pagas {cp_match.group(1)}")
            continue

        months_only_match = re.fullmatch(r"((?:0[1-9]|1[0-2])(?:\s+a\s+(?:0[1-9]|1[0-2])|(?:/(?:0[1-9]|1[0-2]))*)?)", part)
        if months_only_match:
            formatted.append(
                "o extrato, em PDF e OFX, e o relatório de contas pagas "
                f"dos meses {months_for_client_message(months_only_match.group(1))}"
            )
            continue

        statement_match = re.fullmatch(r"([A-Za-z][A-Za-z0-9 ]*?) (?:(PDF|OFX) )?((?:0[1-9]|1[0-2])(?:\s+a\s+(?:0[1-9]|1[0-2])|(?:/(?:0[1-9]|1[0-2]))*)?)", part)
        if statement_match:
            bank = statement_match.group(1).strip()
            file_type = statement_match.group(2)
            months = statement_match.group(3)
            type_text = file_type if file_type else "PDF e OFX"
            formatted.append(f"o extrato {bank}, em {type_text} {month_word_for_client_message(months)} {months_for_client_message(months)}")
            continue

        formatted.append(part)

    if not formatted:
        return faltando
    if len(formatted) == 1:
        return formatted[0]
    return ", ".join(formatted[:-1]) + " e " + formatted[-1]


def months_for_client_message(months: str) -> str:
    months = months.strip()
    if "/" not in months:
        return months

    parts = [part.strip() for part in months.split("/") if part.strip()]
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        return f"{parts[0]} e {parts[1]}"
    return ", ".join(parts[:-1]) + f" e {parts[-1]}"


def month_word_for_client_message(months: str) -> str:
    text = months.strip()
    if "/" in text:
        parts = [part.strip() for part in text.split("/") if part.strip()]
        return "dos meses" if len(parts) > 1 else "do mês"
    if re.fullmatch(r"0[1-9]|1[0-2]", text):
        return "do mês"
    return "dos meses"


def build_email_contact_map(
    sheet: Worksheet,
    columns: Columns,
    company_folders: list[Path],
) -> dict[str, dict[str, object]]:
    contacts: dict[str, dict[str, object]] = {}
    for row in range(columns.header_row + 1, sheet.max_row + 1):
        codigo = sheet.cell(row=row, column=columns.codigo).value
        empresa = sheet.cell(row=row, column=columns.empresa).value
        if codigo is None and empresa is None:
            continue

        emails = extract_emails(row_values(sheet, row))
        if not emails:
            continue

        folder = find_company_folder(company_folders, codigo, empresa)
        for email_address in emails:
            contacts[email_address.casefold()] = {
                "codigo": codigo,
                "empresa": empresa,
                "pasta": folder,
            }
    return contacts


def build_whatsapp_scan_targets(
    sheet: Worksheet,
    columns: Columns,
    company_folders: list[Path],
) -> list[dict[str, object]]:
    via_column = find_optional_column(sheet, columns.header_row, {"via", "canal"})
    targets: list[dict[str, object]] = []

    for row in range(columns.header_row + 1, sheet.max_row + 1):
        codigo = sheet.cell(row=row, column=columns.codigo).value
        empresa = sheet.cell(row=row, column=columns.empresa).value
        if codigo is None and empresa is None:
            continue

        values = row_values(sheet, row)
        phones = extract_phones(values)
        via = sheet.cell(row=row, column=via_column).value if via_column else None
        via_text = normalize_for_match(via)
        has_whatsapp = "whats" in via_text or bool(phones)
        if not has_whatsapp:
            continue

        codigo_text = re.sub(r"\.0$", "", str(codigo).strip()) if codigo is not None else ""
        targets.append(
            {
                "codigo": codigo_text,
                "empresa": "" if empresa is None else str(empresa),
                "telefone": phones[0] if phones else "",
                "pasta": find_company_folder(company_folders, codigo, empresa),
            }
        )

    return targets


def build_company_records(
    sheet: Worksheet,
    columns: Columns,
    company_folders: list[Path],
) -> list[CompanyRecord]:
    records = []
    for row in range(columns.header_row + 1, sheet.max_row + 1):
        codigo = sheet.cell(row=row, column=columns.codigo).value
        empresa = sheet.cell(row=row, column=columns.empresa).value
        if codigo is None and empresa is None:
            continue

        codigo_text = re.sub(r"\.0$", "", str(codigo).strip()) if codigo is not None else ""
        empresa_text = normalize_for_match(empresa)
        tokens = meaningful_company_tokens(empresa_text)
        records.append(
            CompanyRecord(
                codigo=codigo,
                empresa=empresa,
                pasta=find_company_folder(company_folders, codigo, empresa),
                codigo_text=codigo_text,
                empresa_text=empresa_text,
                empresa_tokens=tuple(tokens),
            )
        )
    return records


def meaningful_company_tokens(company_text: str) -> list[str]:
    ignored = {
        "a",
        "de",
        "da",
        "das",
        "do",
        "dos",
        "e",
        "ltda",
        "me",
        "epp",
        "comercio",
        "servicos",
        "empresa",
        "ver",
        "simples",
        "nacional",
    }
    return [token for token in company_text.split() if len(token) >= 3 and token not in ignored]


def month_from_attachment_name(file_name: str, months: list[int], year: int, default_month: int) -> int:
    found = months_from_file_name(file_name, year, set(months))
    if found:
        return min(found)

    normalized_name = normalize_for_match(file_name)
    for month in months:
        if re.search(rf"\b0?{month}\b", normalized_name):
            return month

    return default_month


def ensure_month_folder(company_folder: Path, month: int, year: int) -> Path:
    folder = month_folder(company_folder, month, year)
    if folder:
        return folder

    folder = company_folder / f"{month:02d} - {year}"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def safe_file_name(file_name: str) -> str:
    name = Path(file_name).name.strip()
    name = re.sub(r'[<>:"/\\|?*]+', " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name or "anexo"


def unique_path(folder: Path, file_name: str) -> Path:
    target = folder / safe_file_name(file_name)
    if not target.exists():
        return target

    stem = target.stem
    suffix = target.suffix
    counter = 1
    while True:
        candidate = folder / f"{stem} ({counter}){suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def is_supported_attachment(file_name: str) -> bool:
    return Path(file_name).suffix.lower() in ATTACHMENT_EXTENSIONS


def message_attachments(message: Message) -> Iterable[tuple[str, bytes]]:
    for part in message.walk():
        if part.is_multipart():
            continue
        file_name = part.get_filename()
        if not file_name or not is_supported_attachment(file_name):
            continue
        payload = part.get_payload(decode=True)
        if payload:
            yield file_name, payload


def attachment_search_text(file_name: str, payload: bytes) -> str:
    parts = [file_name]
    suffix = Path(file_name).suffix.lower()

    if suffix == ".pdf":
        text = extract_pdf_text_from_bytes(payload, max_pages=2)
        if text:
            parts.append(text)
    elif suffix in {".ofx", ".csv", ".txt"}:
        parts.append(decode_text_payload(payload))

    return normalize_for_match("\n".join(parts))


def decode_text_payload(payload: bytes) -> str:
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return payload.decode(encoding, errors="ignore")
        except Exception:
            continue
    return ""


def extract_pdf_text_from_bytes(payload: bytes, max_pages: int = 2) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            return ""

    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            reader = PdfReader(io.BytesIO(payload))
            pages = reader.pages[:max_pages]
            return "\n".join(page.extract_text() or "" for page in pages)
    except Exception as error:
        print(f"Nao consegui ler o PDF recebido por e-mail: {error}")
        return ""


def attachment_company_matches(search_text: str, company: CompanyRecord) -> bool:
    if company.codigo_text and re.search(rf"\b{re.escape(company.codigo_text)}\b", search_text):
        return True

    if company.empresa_text and company.empresa_text in search_text:
        return True

    tokens = [token for token in company.empresa_tokens if token in search_text]
    required = min(2, len(company.empresa_tokens))
    return required > 0 and len(tokens) >= required


def attachment_file_name_matches(file_name_text: str, company: CompanyRecord) -> bool:
    if company.codigo_text and re.search(rf"\b{re.escape(company.codigo_text)}\b", file_name_text):
        return True

    if company.empresa_text and company.empresa_text in file_name_text:
        return True

    return any(len(token) >= 5 and token in file_name_text for token in company.empresa_tokens)


def detect_attachment_company(
    file_name: str,
    payload: bytes,
    companies: list[CompanyRecord],
) -> tuple[CompanyRecord | None, str]:
    file_name_text = normalize_for_match(file_name)
    file_name_matches = [
        company for company in companies if attachment_file_name_matches(file_name_text, company)
    ]
    if len(file_name_matches) == 1:
        return file_name_matches[0], "nome do arquivo"
    if len(file_name_matches) > 1:
        names = ", ".join(str(company.empresa) for company in file_name_matches[:5])
        return None, f"ambiguidade no nome: {names}"

    search_text = attachment_search_text(file_name, payload)
    matches = [company for company in companies if attachment_company_matches(search_text, company)]

    if len(matches) == 1:
        return matches[0], "conteudo do arquivo"

    if len(matches) > 1:
        names = ", ".join(str(company.empresa) for company in matches[:5])
        return None, f"ambiguidade no conteudo: {names}"
    return None, "empresa da planilha nao identificada"


def check_email_attachments(
    email_config: EmailConfig,
    spreadsheet: Path,
    base_folder: Path,
    year: int,
    months: list[int],
    sheet_name: str | None,
    header_row: int | None,
    default_month: int,
    dry_run: bool,
    max_emails: int,
) -> int:
    workbook = load_workbook(spreadsheet, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    columns = find_columns(sheet, header_row)
    company_folders = list_company_folders(base_folder)
    contacts = build_email_contact_map(sheet, columns, company_folders)
    companies = build_company_records(sheet, columns, company_folders)

    print(f"Checando e-mails nao lidos em {email_config.pasta_entrada}...")
    print(f"Contatos de e-mail mapeados: {len(contacts)}")
    saved = 0
    mailbox = imaplib.IMAP4_SSL(email_config.imap_host, email_config.imap_port)
    try:
        mailbox.login(email_config.usuario, email_config.senha)
        mailbox.select(email_config.pasta_entrada)
        status, data = mailbox.uid("search", None, "UNSEEN")
        if status != "OK":
            raise RuntimeError("Nao consegui buscar e-mails nao lidos.")

        message_ids = data[0].split()
        if max_emails > 0:
            message_ids = message_ids[-max_emails:]

        for message_id in message_ids:
            try:
                uid_text = message_id.decode("ascii", errors="ignore")
                status, message_data = mailbox.uid("fetch", message_id, "(RFC822)")
                if status != "OK" or not message_data or not isinstance(message_data[0], tuple):
                    continue

                message = BytesParser(policy=policy.default).parsebytes(message_data[0][1])
                sender = parseaddr(message.get("From", ""))[1].casefold()
                contact = contacts.get(sender)
                if not contact:
                    continue

                summary: dict[str, object] = {
                    "assunto": message.get("Subject", "(sem assunto)"),
                    "remetente": sender,
                    "uid": uid_text,
                    "link": email_message_link(email_config, uid_text),
                    "salvos": [],
                    "ignorados": [],
                    "dry_run": dry_run,
                }

                for file_name, payload in message_attachments(message):
                    company, reason = detect_attachment_company(file_name, payload, companies)
                    if not company:
                        summary["ignorados"].append(f"{file_name} ({reason})")
                        continue

                    company_folder = company.pasta
                    if not isinstance(company_folder, Path):
                        summary["ignorados"].append(f"{file_name} (pasta nao encontrada para {company.empresa})")
                        continue

                    attachment_month = month_from_attachment_name(file_name, months, year, default_month)
                    destination_folder = ensure_month_folder(company_folder, attachment_month, year)
                    destination = unique_path(destination_folder, file_name)
                    if not dry_run:
                        destination.write_bytes(payload)
                    saved += 1
                    summary["salvos"].append(
                        f"{file_name} -> {company.codigo} - {company.empresa} -> {destination} ({reason})"
                    )

                if email_config.marcar_como_lido and not dry_run:
                    mailbox.uid("store", message_id, "+FLAGS", "\\Seen")
                print_email_summary(summary)
            except imaplib.IMAP4.abort as error:
                print(f"Mensagem ignorada por queda de conexao IMAP: {error}")
                break
            except Exception as error:
                print(f"Mensagem ignorada por erro: {error}")
                continue
    finally:
        try:
            mailbox.logout()
        except Exception:
            pass

    return saved


def import_whatsapp_folder(
    source_folder: Path,
    spreadsheet: Path,
    base_folder: Path,
    year: int,
    months: list[int],
    sheet_name: str | None,
    header_row: int | None,
    default_month: int,
    dry_run: bool,
) -> int:
    if not source_folder.exists() or not source_folder.is_dir():
        raise FileNotFoundError(f"Pasta de entrada do WhatsApp nao encontrada: {source_folder}")

    workbook = load_workbook(spreadsheet, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    columns = find_columns(sheet, header_row)
    company_folders = list_company_folders(base_folder)
    companies = build_company_records(sheet, columns, company_folders)

    print(f"Lendo entrada do WhatsApp: {source_folder}")
    print(f"Empresas mapeadas na planilha: {len(companies)}")
    saved = 0
    ignored: list[str] = []
    saved_items: list[str] = []
    for file in sorted(source_folder.iterdir()):
        if not file.is_file():
            continue
        if file.suffix.lower() not in ATTACHMENT_EXTENSIONS:
            ignored.append(f"{file.name} (extensao ignorada)")
            continue

        try:
            payload = file.read_bytes()
        except OSError as error:
            ignored.append(f"{file.name} (erro ao ler: {error})")
            continue

        company, reason = detect_attachment_company(file.name, payload, companies)
        if not company:
            ignored.append(f"{file.name} ({reason})")
            continue

        company_folder = company.pasta
        if not isinstance(company_folder, Path):
            ignored.append(f"{file.name} (pasta nao encontrada para {company.empresa})")
            continue

        attachment_month = month_from_attachment_name(file.name, months, year, default_month)
        destination_folder = ensure_month_folder(company_folder, attachment_month, year)
        destination = unique_path(destination_folder, file.name)
        if not dry_run:
            shutil.move(str(file), str(destination))
        saved += 1
        saved_items.append(f"{file.name} -> {company.codigo} - {company.empresa} -> {destination} ({reason})")

    print("")
    print("=" * 72)
    print(f"Entrada WhatsApp: {source_folder}")
    if saved_items:
        print("Arquivos que seriam importados:" if dry_run else "Arquivos importados:")
        for item in saved_items:
            print(f"  OK {item}")
    if ignored:
        print("Arquivos ignorados:")
        for item in ignored:
            print(f"  - {item}")
    if not saved_items and not ignored:
        print("Nenhum arquivo encontrado para importar.")

    return saved


def email_message_link(email_config: EmailConfig, uid: str) -> str:
    folder = email_config.pasta_entrada.replace(" ", "%20")
    return f"imap://{email_config.usuario}@{email_config.imap_host}/{folder};UID={uid}"


def print_email_summary(summary: dict[str, object]) -> None:
    saved_items = summary["salvos"]
    ignored_items = summary["ignorados"]
    if not saved_items and not ignored_items:
        return

    print("")
    print("=" * 72)
    print(f"E-mail: {summary['assunto']}")
    print(f"Remetente: {summary['remetente']}")
    print(f"ID: {summary['uid']}")
    print(f"Link: {summary['link']}")

    if saved_items:
        print("Anexos que seriam salvos:" if summary["dry_run"] else "Anexos salvos:")
        for item in saved_items:
            print(f"  OK {item}")

    if ignored_items:
        print("Anexos ignorados:")
        for item in ignored_items:
            print(f"  - {item}")


def read_charges(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file, delimiter=";"))


def send_email_charges(email_config: EmailConfig, charges_path: Path, dry_run: bool) -> int:
    rows = read_charges(charges_path)
    email_rows = [
        row
        for row in rows
        if row.get("canal") == "email" and row.get("destino") and row.get("mensagem")
    ]

    print(f"Cobrancas por e-mail encontradas: {len(email_rows)}")
    sent = 0
    context = ssl.create_default_context()
    smtp = None
    if not dry_run:
        smtp = smtplib.SMTP(email_config.smtp_host, email_config.smtp_port)
        smtp.starttls(context=context)
        smtp.login(email_config.usuario, email_config.senha)

    try:
        for row in email_rows:
            message = EmailMessage()
            message["From"] = email_config.usuario
            message["To"] = row["destino"]
            message["Subject"] = "Documentos pendentes"
            message.set_content(row["mensagem"])

            if dry_run:
                print(f"E-mail simulado: {row['empresa']} -> {row['destino']}")
            else:
                assert smtp is not None
                smtp.send_message(message)
                print(f"E-mail enviado: {row['empresa']} -> {row['destino']}")
            sent += 1
    finally:
        if smtp:
            smtp.quit()

    return sent


def normalize_whatsapp_phone(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    digits = digits.lstrip("0")
    if len(digits) in {10, 11}:
        digits = "55" + digits
    return digits


def edit_whatsapp_message(current_message: str) -> str:
    print("")
    print("Digite a nova mensagem. Para finalizar, digite apenas um ponto em uma linha.")
    print("Se quiser manter a mensagem atual, digite apenas ponto.")
    lines = []
    while True:
        line = input()
        if line.strip() == ".":
            break
        lines.append(line)
    new_message = "\n".join(lines).strip()
    return new_message or current_message


def whatsapp_review_action(row: dict[str, str]) -> tuple[str, str]:
    message = row.get("mensagem", "")
    while True:
        print("")
        print("=" * 72)
        print(f"Empresa: {row.get('codigo', '')} - {row.get('empresa', '')}")
        print(f"Destino: {row.get('destino', '')}")
        print("Faltando:")
        print(row.get("faltando", ""))
        print("")
        print("Mensagem:")
        print(message)
        print("")
        action = input("[S] enviar  [E] editar  [P] pular  [Q] sair: ").strip().casefold()
        if action in {"s", "sim", "enviar"}:
            return "send", message
        if action in {"e", "editar"}:
            message = edit_whatsapp_message(message)
            continue
        if action in {"p", "pular"}:
            return "skip", message
        if action in {"q", "sair"}:
            return "quit", message
        print("Opcao invalida.")


def wait_for_whatsapp_ready(page: object) -> None:
    page.goto("https://web.whatsapp.com/", wait_until="domcontentloaded", timeout=60000)
    print("")
    print("WhatsApp Web aberto.")
    print("Se aparecer QR Code, escaneie com o WhatsApp Business do celular.")
    input("Quando o WhatsApp Web estiver carregado, pressione ENTER aqui para continuar...")


def open_whatsapp_login(profile_path: Path) -> None:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as error:
        raise RuntimeError(
            "Playwright nao esta instalado. Rode:\n"
            "  pip install playwright\n"
            "  python -m playwright install chromium"
        ) from error

    profile_path.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            str(profile_path),
            headless=False,
            viewport={"width": 1280, "height": 900},
        )
        page = context.pages[0] if context.pages else context.new_page()
        try:
            wait_for_whatsapp_ready(page)
            print(f"Login do WhatsApp salvo em: {profile_path}")
        finally:
            context.close()


def send_whatsapp_message_to_search_contact(profile_path: Path, contact_name: str, message: str) -> None:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as error:
        raise RuntimeError(
            "Playwright nao esta instalado. Rode:\n"
            "  pip install playwright\n"
            "  python -m playwright install chromium"
        ) from error

    profile_path.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            str(profile_path),
            headless=False,
            viewport={"width": 1280, "height": 900},
        )
        page = context.pages[0] if context.pages else context.new_page()
        try:
            print(f"Abrindo WhatsApp Web para contato: {contact_name}")
            open_whatsapp_chat_by_search(page, contact_name)
            print("Contato aberto. Escrevendo mensagem...")
            fill_whatsapp_message(page, message)
            print("Mensagem escrita. Enviando...")
            click_whatsapp_send(page)
            print(f"Mensagem enviada para aprovacao: {contact_name}")
        finally:
            context.close()


def click_whatsapp_send(page: object) -> None:
    selectors = [
        "button[aria-label='Enviar']",
        "button[aria-label='Send']",
        "span[data-icon='send']",
    ]
    last_error: Exception | None = None
    for selector in selectors:
        try:
            locator = page.locator(selector).last
            locator.click(timeout=15000)
            return
        except Exception as error:
            last_error = error
            continue
    try:
        page.keyboard.press("Enter")
    except Exception as error:
        raise RuntimeError(f"Nao consegui clicar no botao de enviar do WhatsApp: {last_error or error}") from error


def first_working_locator(page: object, selectors: list[str], timeout: int = 10000) -> object:
    last_error: Exception | None = None
    per_selector_timeout = max(1000, min(3000, timeout // max(1, len(selectors))))
    for selector in selectors:
        try:
            locator = page.locator(selector).last
            locator.wait_for(state="visible", timeout=per_selector_timeout)
            return locator
        except Exception as error:
            last_error = error
            continue
    raise RuntimeError(f"Nao encontrei o campo no WhatsApp Web: {last_error}")


def open_whatsapp_chat_by_search(page: object, search_text: str) -> None:
    print("Carregando WhatsApp Web...")
    page.goto("https://web.whatsapp.com/", wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(8000)
    print("Procurando campo de busca do WhatsApp...")
    try:
        search_box = first_working_locator(
            page,
            [
                "input[placeholder='Search or start a new chat']",
                "input[placeholder*='Search']",
                "input[placeholder*='Pesquisar']",
                "input[aria-label*='Search']",
                "input[aria-label*='Pesquisar']",
                "div[contenteditable='true'][data-tab='3']",
                "div[contenteditable='true'][data-tab='2']",
                "div[contenteditable='true'][aria-label*='Pesquisar']",
                "div[contenteditable='true'][aria-label*='Search']",
                "div[aria-label='Caixa de texto de pesquisa'][contenteditable='true']",
                "div[aria-label='Search input textbox'][contenteditable='true']",
                "div[title='Caixa de texto de pesquisa']",
                "div[title='Search input textbox']",
                "div[role='textbox'][contenteditable='true']",
                "div[contenteditable='true']",
            ],
            timeout=30000,
        )
    except Exception as error:
        screenshot = SCRIPT_DIR / "whatsapp_debug.png"
        try:
            page.screenshot(path=str(screenshot), full_page=True)
            print(f"Print da tela do WhatsApp salvo em: {screenshot}")
        except Exception:
            pass
        raise error
    search_box.click()
    print(f"Buscando contato: {search_text}")
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    page.keyboard.insert_text(search_text)
    page.wait_for_timeout(2500)
    try:
        page.get_by_text(search_text, exact=True).first.click(timeout=5000)
    except Exception:
        try:
            page.locator(f"span[title='{search_text}']").first.click(timeout=5000)
        except Exception:
            page.keyboard.press("Enter")
    page.wait_for_timeout(2500)
    print("Chat selecionado.")


def fill_whatsapp_message(page: object, message: str) -> None:
    message_box = first_working_locator(
        page,
        [
            "div[contenteditable='true'][data-tab='10']",
            "div[contenteditable='true'][data-tab='9']",
            "div[contenteditable='true'][aria-label*='Digite uma mensagem']",
            "div[contenteditable='true'][aria-label*='Mensagem']",
            "div[contenteditable='true'][aria-label*='Type a message']",
            "div[contenteditable='true'][aria-placeholder*='Type a message']",
            "div[contenteditable='true'][aria-placeholder*='Digite uma mensagem']",
            "footer div[role='textbox'][contenteditable='true']",
            "footer div[contenteditable='true']",
            "div[role='textbox'][contenteditable='true']",
        ],
        timeout=20000,
    )
    message_box.click()
    page.keyboard.insert_text(message)


def visible_attachment_names(page: object) -> list[str]:
    try:
        text = page.locator("body").inner_text(timeout=10000)
    except Exception as error:
        print(f"Nao consegui ler os anexos visiveis no WhatsApp Web: {error}")
        return []

    pattern = re.compile(
        r"[\wÀ-ÿ\.\-\(\) ]{2,}\.(?:pdf|ofx|xlsx|xls|csv|zip|png|jpg|jpeg)",
        flags=re.IGNORECASE,
    )
    names = []
    for match in pattern.finditer(text):
        name = re.sub(r"\s+", " ", match.group(0)).strip(" .-")
        if name:
            names.append(name)
    return sorted(set(names), key=normalize_for_match)


def saved_file_keys(company_folder: Path | None) -> set[str]:
    if not company_folder or not company_folder.exists():
        return set()
    keys = set()
    for file in company_folder.rglob("*"):
        if file.is_file():
            keys.add(normalize_for_match(file.name))
    return keys


def months_from_visible_attachment_name(file_name: str, year: int) -> set[int]:
    months = months_from_file_name(file_name, year, set(range(1, 13)))
    if months:
        return months

    text = normalize_for_match(file_name)
    month_names = {
        1: {"janeiro", "jan"},
        2: {"fevereiro", "fev"},
        3: {"marco", "mar o", "mar oo", "mar"},
        4: {"abril", "abr"},
        5: {"maio", "mai"},
        6: {"junho", "jun"},
        7: {"julho", "jul"},
        8: {"agosto", "ago"},
        9: {"setembro", "set"},
        10: {"outubro", "out"},
        11: {"novembro", "nov"},
        12: {"dezembro", "dez"},
    }
    found = set()
    padded = f" {text} "
    for month, names in month_names.items():
        if any(f" {name} " in padded for name in names):
            found.add(month)
    return found


def month_folders_to_check(company_folder: Path, file_name: str, year: int) -> list[Path]:
    months = months_from_visible_attachment_name(file_name, year)
    folders = []
    if months:
        for month in sorted(months):
            folder = month_folder(company_folder, month, year)
            if folder:
                folders.append(folder)
        return folders

    return [
        child
        for child in company_folder.iterdir()
        if child.is_dir() and (re.match(r"^0?[1-9]|1[0-2]", normalize(child.name)) or str(year) in child.name)
    ]


def has_any_statement_file(month_path: Path, extension: str) -> bool:
    for file in month_path.iterdir():
        if file.is_file() and file.suffix.lower() == extension and bank_from_statement_file(file):
            return True
    return False


def attachment_equivalent_saved(file_name: str, company_folder: Path | None, year: int) -> bool:
    if not company_folder or not company_folder.exists():
        return False

    if normalize_for_match(file_name) in saved_file_keys(company_folder):
        return True

    suffix = Path(file_name).suffix.lower()
    text = normalize_for_match(Path(file_name).stem)
    compact_text = text.replace(" ", "")
    folders = month_folders_to_check(company_folder, file_name, year)
    if not folders:
        return False

    is_cp = (
        "contas pagas" in text
        or "conta paga" in text
        or "contas pg" in text
        or "conta pg" in text
        or "contaspagas" in compact_text
        or "contapaga" in compact_text
        or "contaspg" in compact_text
        or "contapg" in compact_text
    )
    if is_cp:
        return all(has_paid_accounts_report(folder) for folder in folders)

    if suffix not in STATEMENT_EXTENSIONS:
        return False

    bank = known_bank_in_file_name(text)
    looks_like_statement = (
        bank is not None
        or "extrato" in text
        or re.search(r"\bextr\b", text) is not None
        or "bancario" in text
        or "conta corrente" in text
    )
    if not looks_like_statement:
        return False

    if bank:
        return all(has_statement_file(folder, bank, suffix) for folder in folders)
    return all(has_any_statement_file(folder, suffix) for folder in folders)


def unsaved_visible_attachments(page: object, company_folder: Path | None, year: int) -> list[str]:
    saved = saved_file_keys(company_folder)
    unsaved = []
    for name in visible_attachment_names(page):
        if normalize_for_match(name) in saved:
            continue
        if not attachment_equivalent_saved(name, company_folder, year):
            unsaved.append(name)
    return unsaved


def open_whatsapp_chat_by_phone_or_code(page: object, phone: str, code: str) -> str:
    normalized_phone = normalize_whatsapp_phone(phone)
    if normalized_phone:
        page.goto(f"https://web.whatsapp.com/send?phone={normalized_phone}", wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(5000)
        return normalized_phone

    if not code:
        raise ValueError("Contato sem telefone e sem codigo.")

    open_whatsapp_chat_by_search(page, code)
    return f"codigo {code}"


def check_whatsapp_documents(
    spreadsheet: Path,
    base_folder: Path,
    year: int,
    sheet_name: str | None,
    header_row: int | None,
    profile_path: Path,
    limit: int,
) -> int:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as error:
        raise RuntimeError(
            "Playwright nao esta instalado. Rode:\n"
            "  pip install playwright\n"
            "  python -m playwright install chromium"
        ) from error

    workbook = load_workbook(spreadsheet, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    columns = find_columns(sheet, header_row)
    company_folders = list_company_folders(base_folder)
    targets = build_whatsapp_scan_targets(sheet, columns, company_folders)
    if limit > 0:
        targets = targets[:limit]

    print(f"Empresas/contatos para checar no WhatsApp: {len(targets)}")
    print("Obs.: esta checagem ve nomes de anexos visiveis no WhatsApp Web; ela avisa para voce conferir.")

    found = 0
    profile_path.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            str(profile_path),
            headless=False,
            viewport={"width": 1280, "height": 900},
        )
        page = context.pages[0] if context.pages else context.new_page()
        try:
            wait_for_whatsapp_ready(page)
            for index, target in enumerate(targets, start=1):
                codigo = str(target["codigo"])
                empresa = str(target["empresa"])
                phone = str(target["telefone"])
                company_folder = target["pasta"] if isinstance(target["pasta"], Path) else None

                print("")
                print("-" * 72)
                print(f"[{index}/{len(targets)}] Checando {codigo} - {empresa}")
                try:
                    opened_by = open_whatsapp_chat_by_phone_or_code(page, phone, codigo)
                    candidates = unsaved_visible_attachments(page, company_folder, year)
                except Exception as error:
                    print(f"  Nao consegui checar este contato: {error}")
                    continue

                if not candidates:
                    print(f"  Nenhum anexo visivel nao salvo encontrado ({opened_by}).")
                    continue

                found += len(candidates)
                print(f"  Possiveis documentos nao salvos ({opened_by}):")
                for name in candidates[:20]:
                    print(f"    - {name}")
                if len(candidates) > 20:
                    print(f"    ... e mais {len(candidates) - 20}")
        finally:
            context.close()

    return found


def send_whatsapp_charges(charges_path: Path, profile_path: Path, dry_run: bool) -> int:
    rows = read_charges(charges_path)
    whatsapp_all_rows = [
        row
        for row in rows
        if row.get("canal") == "whatsapp"
    ]
    whatsapp_rows = [
        row
        for row in whatsapp_all_rows
        if row.get("canal") == "whatsapp" and row.get("mensagem") and (row.get("destino") or row.get("codigo"))
    ]
    without_route = [row for row in whatsapp_all_rows if not row.get("destino") and not row.get("codigo")]
    by_code = [row for row in whatsapp_rows if not row.get("destino") and row.get("codigo")]

    if not whatsapp_rows:
        print("Nenhuma cobranca por WhatsApp pronta para envio.")
        if without_route:
            print(f"Cobrancas WhatsApp sem telefone e sem codigo: {len(without_route)}")
            for row in without_route[:10]:
                print(f"  - {row.get('codigo', '')} - {row.get('empresa', '')}")
        return 0

    print(f"Cobrancas por WhatsApp prontas para envio: {len(whatsapp_rows)}")
    if by_code:
        print(f"Cobrancas sem telefone que serao buscadas pelo codigo: {len(by_code)}")
    if without_route:
        print(f"Cobrancas WhatsApp sem telefone e sem codigo ignoradas: {len(without_route)}")
    print(f"Arquivo de cobrancas: {charges_path}")
    print(f"Perfil do WhatsApp Web: {profile_path}")
    if dry_run:
        sent = 0
        for row in whatsapp_rows:
            action, _message = whatsapp_review_action(row)
            if action == "quit":
                break
            if action == "send":
                print(f"WhatsApp simulado: {row.get('empresa', '')} -> {row.get('destino', '')}")
                sent += 1
        return sent

    try:
        from playwright.sync_api import sync_playwright
    except ImportError as error:
        raise RuntimeError(
            "Playwright nao esta instalado. Rode:\n"
            "  pip install playwright\n"
            "  python -m playwright install chromium"
        ) from error

    sent = 0
    profile_path.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            str(profile_path),
            headless=False,
            viewport={"width": 1280, "height": 900},
        )
        page = context.pages[0] if context.pages else context.new_page()
        try:
            wait_for_whatsapp_ready(page)
            for row in whatsapp_rows:
                action, message = whatsapp_review_action(row)
                if action == "quit":
                    break
                if action == "skip":
                    print(f"WhatsApp pulado: {row.get('empresa', '')}")
                    continue

                phone = normalize_whatsapp_phone(row.get("destino", ""))
                if phone:
                    encoded_message = urllib.parse.quote(message)
                    page.goto(
                        f"https://web.whatsapp.com/send?phone={phone}&text={encoded_message}",
                        wait_until="domcontentloaded",
                        timeout=60000,
                    )
                    page.wait_for_timeout(5000)
                    destination_text = phone
                else:
                    search_text = str(row.get("codigo", "")).strip()
                    print(f"Sem telefone. Procurando contato pelo codigo: {search_text}")
                    open_whatsapp_chat_by_search(page, search_text)
                    fill_whatsapp_message(page, message)
                    destination_text = f"codigo {search_text}"

                click_whatsapp_send(page)
                print(f"WhatsApp enviado: {row.get('empresa', '')} -> {destination_text}")
                sent += 1
                page.wait_for_timeout(1500)
        finally:
            context.close()

    return sent


def build_approval_message(charges_path: Path, max_items: int = 12) -> str | None:
    if not charges_path.exists():
        return None

    rows = [
        row
        for row in read_charges(charges_path)
        if row.get("mensagem")
        and row.get("faltando")
        and row.get("faltando") != "-"
        and row.get("canal") != "sem contato"
        and row.get("destino")
    ]
    if not rows:
        return None

    lines = [
        "Arthur, encontrei cobrancas pendentes para revisar.",
        "",
        f"Arquivo: {charges_path}",
        "",
    ]
    for index, row in enumerate(rows[:max_items], start=1):
        lines.extend(
            [
                f"{index}. {row.get('codigo', '')} - {row.get('empresa', '')}",
                f"Canal: {row.get('canal', '')} | Destino: {row.get('destino', '') or 'sem destino'}",
                f"Faltando: {row.get('faltando', '')}",
                f"Mensagem: {row.get('mensagem', '')}",
                "",
            ]
        )

    if len(rows) > max_items:
        lines.append(f"... e mais {len(rows) - max_items} cobrancas no CSV.")
        lines.append("")

    lines.extend(
        [
            "Revise o CSV e depois rode:",
            f'python bot.py --enviar-whatsapp --cobrancas "{charges_path}"',
        ]
    )
    return "\n".join(lines)


def dummy_email_config() -> EmailConfig:
    return EmailConfig(
        imap_host="",
        imap_port=993,
        smtp_host="",
        smtp_port=587,
        usuario="simulacao@example.com",
        senha="",
        pasta_entrada="INBOX",
        marcar_como_lido=False,
    )


def missing_text(missing: list[str]) -> str:
    if not missing:
        return "-"

    cp_months: list[int] = []
    month_only: list[int] = []
    statement_missing: dict[str, dict[str, list[int]]] = {}
    others = []

    for item in missing:
        month_match = re.fullmatch(r"0[1-9]|1[0-2]", item)
        if month_match:
            month_only.append(int(item))
            continue

        match = re.fullmatch(r"(.+?) - (0[1-9]|1[0-2])", item)
        if not match:
            others.append(item)
            continue

        label = match.group(1)
        month = int(match.group(2))

        if label == "CP":
            cp_months.append(month)
            continue

        statement_match = re.fullmatch(r"(.+?) (PDF|OFX)", label)
        if statement_match:
            bank = statement_match.group(1).upper()
            file_type = statement_match.group(2)
            statement_missing.setdefault(bank, {"PDF": [], "OFX": []})[file_type].append(month)
            continue

        others.append(item)

    formatted = []
    if cp_months:
        formatted.append(f"CP {format_months(cp_months)}")
    if month_only:
        formatted.append(format_months(month_only))

    for bank in sorted(statement_missing):
        pdf_months = set(statement_missing[bank]["PDF"])
        ofx_months = set(statement_missing[bank]["OFX"])
        both_missing = sorted(pdf_months & ofx_months)
        only_pdf = sorted(pdf_months - ofx_months)
        only_ofx = sorted(ofx_months - pdf_months)

        if both_missing:
            formatted.append(f"{bank} {format_months(both_missing)}")
        if only_pdf:
            formatted.append(f"{bank} PDF {format_months(only_pdf)}")
        if only_ofx:
            formatted.append(f"{bank} OFX {format_months(only_ofx)}")

    formatted.extend(others)
    return ", ".join(formatted)


def format_months(months: list[int]) -> str:
    ordered = sorted(set(months))
    if not ordered:
        return ""

    parts = []
    current_start = ordered[0]
    previous = ordered[0]

    for month in ordered[1:]:
        if month == previous + 1:
            previous = month
            continue
        parts.extend(format_month_range(current_start, previous))
        current_start = previous = month

    parts.extend(format_month_range(current_start, previous))
    return "/".join(parts)


def format_month_range(start: int, end: int) -> list[str]:
    if start == end:
        return [f"{start:02d}"]
    if end == start + 1:
        return [f"{start:02d}", f"{end:02d}"]
    return [f"{start:02d} a {end:02d}"]


def backup_file(path: Path) -> Path:
    backup = path.with_name(f"{path.stem}.backup{path.suffix}")
    counter = 1
    while backup.exists():
        backup = path.with_name(f"{path.stem}.backup-{counter}{path.suffix}")
        counter += 1
    shutil.copy2(path, backup)
    return backup


def update_workbook(
    spreadsheet: Path,
    base_folder: Path,
    year: int,
    months: list[int],
    sheet_name: str | None,
    header_row: int | None,
    dry_run: bool,
    no_backup: bool,
    detailed: bool,
    report_path: Path | None,
    charges_path: Path | None,
) -> tuple[int, Path | None]:
    if spreadsheet.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Use uma planilha .xlsx ou .xlsm. Arquivos .xls antigos nao sao suportados.")

    print(f"Planilha: {spreadsheet}")
    print(f"Pasta base: {base_folder}")
    print(f"Ano: {year} | Meses: {', '.join(month_label(month) for month in months)}")
    workbook = load_workbook(spreadsheet, keep_vba=spreadsheet.suffix.lower() == ".xlsm")
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    columns = find_columns(sheet, header_row)
    via_column = find_optional_column(sheet, columns.header_row, {"via", "canal"})
    company_folders = list_company_folders(base_folder)
    expected = month_folder_names(months, year)
    print(f"Aba analisada: {sheet.title}")
    print(f"Pastas de empresas encontradas: {len(company_folders)}")

    updated = 0
    analyzed = 0
    report_rows: list[dict[str, str]] = []
    charge_rows: list[dict[str, str]] = []
    for row in range(columns.header_row + 1, sheet.max_row + 1):
        codigo = sheet.cell(row=row, column=columns.codigo).value
        empresa = sheet.cell(row=row, column=columns.empresa).value
        if codigo is None and empresa is None:
            continue

        analyzed += 1
        company_folder = find_company_folder(company_folders, codigo, empresa)
        missing = missing_items(company_folder, expected)
        new_value = missing_text(missing)
        folder_text = str(company_folder) if company_folder else ""
        values = row_values(sheet, row)
        emails = extract_emails(values)
        phones = extract_phones(values)
        via = sheet.cell(row=row, column=via_column).value if via_column else None
        channel, destination = choose_charge_channel(via, phones, emails)

        if detailed:
            print(f"{codigo or ''} - {empresa or ''}: {new_value} | {channel}: {destination}")

        if detailed or report_path:
            report_rows.append(
                {
                    "codigo": "" if codigo is None else str(codigo),
                    "empresa": "" if empresa is None else str(empresa),
                    "canal": channel,
                    "telefone": phones[0] if phones else "",
                    "email": emails[0] if emails else "",
                    "pasta": folder_text,
                    "arquivos_analisados": " | ".join(analyzed_files(company_folder, expected)),
                    "faltando": new_value,
                }
            )

        if charges_path and new_value != "-":
            charge_rows.append(
                {
                    "codigo": "" if codigo is None else str(codigo),
                    "empresa": "" if empresa is None else str(empresa),
                    "canal": channel,
                    "destino": destination,
                    "faltando": new_value,
                    "mensagem": charge_message(empresa, new_value, channel),
                }
            )

        cell = sheet.cell(row=row, column=columns.faltando)
        if cell.value != new_value:
            cell.value = new_value
            updated += 1

    print(f"Empresas analisadas: {analyzed}")
    print(f"Celulas com alteracao detectada: {updated}")
    if charges_path:
        print(f"Cobrancas geradas no CSV: {len(charge_rows)}")
    if report_path:
        write_report(report_path, report_rows)
    if charges_path:
        write_charges(charges_path, charge_rows)

    if dry_run:
        return updated, None

    backup = None if no_backup else backup_file(spreadsheet)
    workbook.save(spreadsheet)
    return updated, backup


def assistant_cycle(
    spreadsheet: Path,
    base_folder: Path,
    year: int,
    months: list[int],
    sheet_name: str | None,
    header_row: int | None,
    email_config_path: Path,
    whatsapp_inbox: Path,
    charges_path: Path,
    max_emails: int,
    default_month: int,
    dry_run: bool,
    approval_contact: str | None,
    whatsapp_profile: Path,
) -> None:
    print("")
    print("=" * 72)
    print(f"Assistente iniciado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    if email_config_path.exists():
        try:
            email_config = load_email_config(email_config_path)
            saved = check_email_attachments(
                email_config=email_config,
                spreadsheet=spreadsheet,
                base_folder=base_folder,
                year=year,
                months=months,
                sheet_name=sheet_name,
                header_row=header_row,
                default_month=default_month,
                dry_run=dry_run,
                max_emails=max_emails,
            )
            if saved:
                print(f"Assistente: anexos de e-mail {'que seriam salvos' if dry_run else 'salvos'}: {saved}")
        except Exception as error:
            print(f"Assistente: erro ao checar e-mail: {error}")
    else:
        print(f"Assistente: e-mail ignorado, config nao encontrada: {email_config_path}")

    whatsapp_inbox.mkdir(parents=True, exist_ok=True)
    try:
        imported = import_whatsapp_folder(
            source_folder=whatsapp_inbox,
            spreadsheet=spreadsheet,
            base_folder=base_folder,
            year=year,
            months=months,
            sheet_name=sheet_name,
            header_row=header_row,
            default_month=default_month,
            dry_run=dry_run,
        )
        if imported:
            print(f"Assistente: arquivos do WhatsApp {'que seriam importados' if dry_run else 'importados'}: {imported}")
    except Exception as error:
        print(f"Assistente: erro ao importar entrada do WhatsApp: {error}")

    try:
        updated, backup = update_workbook(
            spreadsheet=spreadsheet,
            base_folder=base_folder,
            year=year,
            months=months,
            sheet_name=sheet_name,
            header_row=header_row,
            dry_run=dry_run,
            no_backup=False,
            detailed=False,
            report_path=None,
            charges_path=charges_path,
        )
        if dry_run:
            print(f"Assistente: {updated} celulas seriam atualizadas.")
        else:
            print(f"Assistente: planilha atualizada, celulas alteradas: {updated}")
            if backup:
                print(f"Assistente: backup criado: {backup}")
        print(f"Assistente: cobrancas para revisao: {charges_path}")
    except PermissionError as error:
        print(f"Assistente: nao consegui salvar. Feche a planilha/CSV no Excel e tente de novo: {error}")
    except Exception as error:
        print(f"Assistente: erro ao atualizar planilha/cobrancas: {error}")

    if approval_contact:
        try:
            approval_message = build_approval_message(charges_path)
            if approval_message:
                if dry_run:
                    print("")
                    print("=" * 72)
                    print(f"Mensagem de aprovacao que seria enviada para {approval_contact}:")
                    print(approval_message)
                else:
                    send_whatsapp_message_to_search_contact(whatsapp_profile, approval_contact, approval_message)
            else:
                print("Assistente: nenhuma cobranca para enviar para aprovacao.")
        except Exception as error:
            print(f"Assistente: erro ao enviar aprovacao por WhatsApp: {error}")

    print(f"Assistente finalizou ciclo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")


def run_assistant(
    args: argparse.Namespace,
    months: list[int],
) -> None:
    default_month = args.mes_recebimento or max(months)
    print("Modo assistente ativo.")
    print(f"Intervalo: {args.intervalo_assistente} segundos")
    print(f"Entrada WhatsApp: {args.entrada_whatsapp}")
    print(f"Cobrancas para revisar: {args.cobrancas or DEFAULT_ASSISTANT_CHARGES}")
    print("Para parar, pressione Ctrl+C.")

    charges_path = args.cobrancas or DEFAULT_ASSISTANT_CHARGES
    while True:
        assistant_cycle(
            spreadsheet=args.planilha,
            base_folder=args.pasta_base,
            year=args.ano,
            months=months,
            sheet_name=args.aba,
            header_row=args.linha_cabecalho,
            email_config_path=args.email_config,
            whatsapp_inbox=args.entrada_whatsapp,
            charges_path=charges_path,
            max_emails=args.limite_emails,
            default_month=default_month,
            dry_run=args.dry_run,
            approval_contact=args.aprovacao_whatsapp,
            whatsapp_profile=args.whatsapp_profile,
        )
        if args.uma_vez:
            break
        time.sleep(max(10, args.intervalo_assistente))


def validate_meta_signature(payload: bytes, signature_header: str | None, app_secret: str | None) -> bool:
    if not app_secret:
        return True
    if not signature_header or not signature_header.startswith("sha256="):
        return False

    expected = hmac.new(app_secret.encode("utf-8"), payload, hashlib.sha256).hexdigest()
    received = signature_header.split("=", 1)[1]
    return hmac.compare_digest(expected, received)


def append_webhook_event(log_path: Path, payload: dict[str, object]) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    event = {
        "received_at": datetime.now().isoformat(timespec="seconds"),
        "payload": payload,
    }
    with log_path.open("a", encoding="utf-8") as file:
        file.write(json.dumps(event, ensure_ascii=False) + "\n")


def whatsapp_webhook_handler(verify_token: str, log_path: Path, app_secret: str | None) -> type[BaseHTTPRequestHandler]:
    class WhatsAppWebhookHandler(BaseHTTPRequestHandler):
        def log_message(self, format: str, *args: object) -> None:
            print(f"Webhook: {self.address_string()} - {format % args}")

        def send_text(self, status: int, text: str) -> None:
            body = text.encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def do_GET(self) -> None:
            parsed = urllib.parse.urlparse(self.path)
            if parsed.path != "/webhook":
                self.send_text(404, "not found")
                return

            params = urllib.parse.parse_qs(parsed.query)
            mode = params.get("hub.mode", [""])[0]
            token = params.get("hub.verify_token", [""])[0]
            challenge = params.get("hub.challenge", [""])[0]

            if mode == "subscribe" and token == verify_token:
                self.send_text(200, challenge)
                print("Webhook verificado pela Meta.")
                return

            self.send_text(403, "forbidden")

        def do_POST(self) -> None:
            parsed = urllib.parse.urlparse(self.path)
            if parsed.path != "/webhook":
                self.send_text(404, "not found")
                return

            length = int(self.headers.get("Content-Length", "0"))
            payload_bytes = self.rfile.read(length)
            signature = self.headers.get("X-Hub-Signature-256")
            if not validate_meta_signature(payload_bytes, signature, app_secret):
                self.send_text(403, "invalid signature")
                return

            try:
                payload = json.loads(payload_bytes.decode("utf-8") or "{}")
            except json.JSONDecodeError:
                self.send_text(400, "invalid json")
                return

            append_webhook_event(log_path, payload)
            print(f"Webhook recebido e salvo: {log_path}")
            self.send_text(200, "ok")

    return WhatsAppWebhookHandler


def run_whatsapp_webhook(host: str, port: int, verify_token: str, log_path: Path, app_secret: str | None) -> None:
    server = ThreadingHTTPServer((host, port), whatsapp_webhook_handler(verify_token, log_path, app_secret))
    print("Webhook WhatsApp ativo.")
    print(f"Local: http://{host}:{port}/webhook")
    print(f"Verify token: {verify_token}")
    print(f"Log de eventos: {log_path}")
    print("Para parar, pressione Ctrl+C.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("Encerrando webhook...")
    finally:
        server.server_close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Atualiza a coluna FALTANDO da planilha conforme as pastas mensais de cada empresa."
    )
    parser.add_argument(
        "--pasta-base",
        default=PASTA_BASE,
        type=Path,
        help=f"Pasta onde ficam as pastas das empresas. Padrao: {PASTA_BASE}",
    )
    parser.add_argument(
        "--planilha",
        default=PLANILHA,
        type=Path,
        help=f"Arquivo .xlsx ou .xlsm que sera atualizado. Padrao: {PLANILHA}",
    )
    parser.add_argument("--ano", type=int, default=ANO, help=f"Ano das pastas mensais. Padrao: {ANO}.")
    parser.add_argument(
        "--meses",
        help="Meses para conferir. Exemplos: '1-4', '1,2,3,4' ou '1 2 3 4'. Padrao: ate o mes anterior.",
    )
    parser.add_argument(
        "--incluir-mes-atual",
        action="store_true",
        help="Quando --meses nao for informado, confere tambem o mes atual.",
    )
    parser.add_argument("--aba", help="Nome da aba. Se nao informar, usa a primeira aba.")
    parser.add_argument("--linha-cabecalho", type=int, help="Linha dos cabecalhos. Se nao informar, procura nas 20 primeiras.")
    parser.add_argument("--dry-run", action="store_true", help="Mostra quantas linhas mudariam, sem salvar.")
    parser.add_argument("--sem-backup", action="store_true", help="Nao cria copia .backup antes de salvar.")
    parser.add_argument("--detalhado", action="store_true", help="Mostra no terminal o resultado de cada empresa.")
    parser.add_argument("--relatorio", type=Path, help="Salva um relatorio CSV com arquivos analisados e faltantes.")
    parser.add_argument("--cobrancas", type=Path, help="Salva um CSV com canal, destino e mensagem de cobranca.")
    parser.add_argument("--email-config", type=Path, default=EMAIL_CONFIG_PATH, help="Arquivo JSON com IMAP/SMTP.")
    parser.add_argument("--checar-email", action="store_true", help="Baixa anexos recebidos dos e-mails da planilha.")
    parser.add_argument("--limite-emails", type=int, default=25, help="Quantidade maxima de e-mails nao lidos para checar.")
    parser.add_argument("--mes-recebimento", type=int, help="Mes usado para anexos cujo nome nao informa o mes.")
    parser.add_argument("--importar-whatsapp", type=Path, help="Importa documentos baixados do WhatsApp de uma pasta de entrada.")
    parser.add_argument("--entrada-whatsapp", type=Path, default=DEFAULT_WHATSAPP_INBOX, help="Pasta monitorada pelo modo assistente para documentos do WhatsApp.")
    parser.add_argument("--enviar-emails", action="store_true", help="Envia por SMTP as cobrancas com canal email.")
    parser.add_argument("--enviar-whatsapp", action="store_true", help="Envia cobrancas com canal whatsapp via WhatsApp Web, com revisao manual.")
    parser.add_argument("--login-whatsapp", action="store_true", help="Abre o WhatsApp Web para fazer login e salvar a sessao.")
    parser.add_argument("--checar-whatsapp", action="store_true", help="Procura no WhatsApp Web possiveis documentos ainda nao salvos.")
    parser.add_argument("--limite-whatsapp", type=int, default=0, help="Limita a quantidade de contatos checados no WhatsApp. 0 checa todos.")
    parser.add_argument("--assistente", action="store_true", help="Roda continuamente checando e-mail, entrada do WhatsApp, planilha e cobrancas.")
    parser.add_argument("--intervalo-assistente", type=int, default=300, help="Intervalo em segundos entre ciclos do assistente.")
    parser.add_argument(
        "--aprovacao-whatsapp",
        nargs="?",
        const=DEFAULT_APPROVAL_CONTACT,
        help=f"Contato/grupo no WhatsApp que recebe o resumo das cobrancas. Padrao: {DEFAULT_APPROVAL_CONTACT}.",
    )
    parser.add_argument("--teste-aprovacao-whatsapp", help="Envia uma mensagem curta de teste para um contato do WhatsApp.")
    parser.add_argument("--webhook-whatsapp", action="store_true", help="Inicia o servidor de webhook da API oficial do WhatsApp.")
    parser.add_argument("--webhook-host", default="0.0.0.0", help="Host/interface do servidor webhook.")
    parser.add_argument("--webhook-port", type=int, default=8000, help="Porta do servidor webhook.")
    parser.add_argument(
        "--webhook-token",
        default=os.environ.get("WHATSAPP_VERIFY_TOKEN", DEFAULT_WEBHOOK_TOKEN),
        help="Token de verificacao usado no painel da Meta.",
    )
    parser.add_argument("--webhook-log", type=Path, default=DEFAULT_WEBHOOK_LOG, help="Arquivo JSONL onde eventos recebidos serao salvos.")
    parser.add_argument(
        "--meta-app-secret",
        default=os.environ.get("META_APP_SECRET", ""),
        help="App Secret da Meta para validar X-Hub-Signature-256. Opcional, mas recomendado em producao.",
    )
    parser.add_argument("--uma-vez", action="store_true", help="No modo assistente, roda apenas um ciclo e encerra.")
    parser.add_argument(
        "--whatsapp-profile",
        type=Path,
        default=SCRIPT_DIR / "whatsapp_profile",
        help="Pasta onde o login do WhatsApp Web fica salvo.",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    months = parse_months(args.meses, args.ano, args.incluir_mes_atual)

    if args.webhook_whatsapp:
        if args.webhook_token == DEFAULT_WEBHOOK_TOKEN:
            print(
                "Aviso: usando o webhook-token padrao (previsivel). "
                "Defina a variavel de ambiente WHATSAPP_VERIFY_TOKEN com um valor proprio."
            )
        if not args.meta_app_secret:
            print(
                "Aviso: META_APP_SECRET nao configurado. "
                "As requisicoes recebidas nao terao a assinatura X-Hub-Signature-256 validada."
            )
        run_whatsapp_webhook(
            host=args.webhook_host,
            port=args.webhook_port,
            verify_token=args.webhook_token,
            log_path=args.webhook_log,
            app_secret=args.meta_app_secret or None,
        )
        return

    if args.teste_aprovacao_whatsapp:
        send_whatsapp_message_to_search_contact(
            args.whatsapp_profile,
            args.teste_aprovacao_whatsapp,
            f"Teste do bot em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        )
        return

    if args.assistente:
        run_assistant(args, months)
        return

    if args.login_whatsapp:
        open_whatsapp_login(args.whatsapp_profile)
        return

    if args.checar_whatsapp:
        found = check_whatsapp_documents(
            spreadsheet=args.planilha,
            base_folder=args.pasta_base,
            year=args.ano,
            sheet_name=args.aba,
            header_row=args.linha_cabecalho,
            profile_path=args.whatsapp_profile,
            limit=args.limite_whatsapp,
        )
        print(f"Possiveis documentos nao salvos encontrados: {found}")
        return

    if args.checar_email:
        email_config = load_email_config(args.email_config)
        default_month = args.mes_recebimento or max(months)
        saved = check_email_attachments(
            email_config=email_config,
            spreadsheet=args.planilha,
            base_folder=args.pasta_base,
            year=args.ano,
            months=months,
            sheet_name=args.aba,
            header_row=args.linha_cabecalho,
            default_month=default_month,
            dry_run=args.dry_run,
            max_emails=args.limite_emails,
        )
        print(f"Anexos {'que seriam salvos' if args.dry_run else 'salvos'}: {saved}")

    if args.importar_whatsapp:
        default_month = args.mes_recebimento or max(months)
        imported = import_whatsapp_folder(
            source_folder=args.importar_whatsapp,
            spreadsheet=args.planilha,
            base_folder=args.pasta_base,
            year=args.ano,
            months=months,
            sheet_name=args.aba,
            header_row=args.linha_cabecalho,
            default_month=default_month,
            dry_run=args.dry_run,
        )
        print(f"Arquivos do WhatsApp {'que seriam importados' if args.dry_run else 'importados'}: {imported}")

    print("Abrindo pasta e planilha...")
    should_generate_charges = bool(args.cobrancas and not args.enviar_emails and not args.enviar_whatsapp)
    updated, backup = update_workbook(
        spreadsheet=args.planilha,
        base_folder=args.pasta_base,
        year=args.ano,
        months=months,
        sheet_name=args.aba,
        header_row=args.linha_cabecalho,
        dry_run=args.dry_run,
        no_backup=args.sem_backup,
        detailed=args.detalhado,
        report_path=args.relatorio,
        charges_path=args.cobrancas if should_generate_charges else None,
    )

    checked = ", ".join(f"{calendar.month_abbr[month]}({month:02d})" for month in months)
    print(f"Meses conferidos: {checked}")
    if args.dry_run:
        print(f"Simulacao concluida: {updated} celulas seriam atualizadas.")
    else:
        print(f"Planilha atualizada: {args.planilha}")
        print(f"Celulas alteradas: {updated}")
        if backup:
            print(f"Backup criado: {backup}")
    if args.relatorio:
        print(f"Relatorio criado: {args.relatorio}")
    if should_generate_charges:
        print(f"Lista de cobrancas criada: {args.cobrancas}")
    if args.enviar_emails:
        if not args.cobrancas:
            raise ValueError("Informe --cobrancas para enviar e-mails.")
        email_config = dummy_email_config() if args.dry_run and not args.email_config.exists() else load_email_config(args.email_config)
        sent = send_email_charges(email_config, args.cobrancas, args.dry_run)
        print(f"E-mails {'que seriam enviados' if args.dry_run else 'enviados'}: {sent}")
    if args.enviar_whatsapp:
        if not args.cobrancas:
            raise ValueError("Informe --cobrancas para enviar WhatsApp.")
        sent = send_whatsapp_charges(args.cobrancas, args.whatsapp_profile, args.dry_run)
        print(f"WhatsApps {'que seriam enviados' if args.dry_run else 'enviados'}: {sent}")


if __name__ == "__main__":
    main()
